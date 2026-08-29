#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline de carga: lê o(s) .xlsx do repositório de dados
(andresilvatorres85-bit/emendas.apresentadas.ploa), filtra o Órgão 52000
(Ministério da Defesa) / setor da Defesa, calcula colunas derivadas e gera o
JSON consumido pelo front-end (public/dados.json).

Uso:
    python3 scripts/processar_dados.py [pasta_com_xlsx]

Se nenhuma pasta for informada, o script baixa os .xlsx direto do GitHub
(listando o conteúdo do repositório via API pública).

Colunas derivadas:
  - ano    : exercício das emendas (ver REGRA 0)
  - cmila  : Comando Militar de Área, deduzido de "Autor (UF)" (ver regra abaixo)
  - inconsistencias : lista de descrições de inconsistência OM x UO (ver regra abaixo)

============================================================================
REGRA 0 — De onde vem o ANO e como os anos são normalizados
============================================================================
O repositório tem dois formatos de arquivo:

  a) arquivo de UM exercício ("PLOA 2026 - EMENDAS APRESENTADAS.xlsx"), com as
     abas "Todos os Setores" e "Setor NN". O ano sai do NOME DO ARQUIVO.
  b) arquivo consolidado ("Historico_emendas_apresentadas.xlsx"), com UMA ABA
     POR ANO ("2026", "2025", …). O ano sai do NOME DA ABA.

PRECEDÊNCIA: quando o mesmo ano aparece nos dois formatos, vence o arquivo do
exercício (a). Ele é o que acompanha a tramitação e é atualizado; a aba do
consolidado é uma fotografia. Em 2026 os dois somam o mesmo valor total
(R$ 4.009.124.231,00) — o consolidado só traz 18 linhas de valor zero a mais.

As planilhas não são idênticas entre si. Estas diferenças são normalizadas:

  - SETOR: em 2023 o setor da Defesa tem "Setor (Cod)" = 08; de 2024 em diante,
    13. O recorte, portanto, é feito pelo NOME do setor ("DEFESA") somado ao
    "Órgão (Cod)" = 52000, e não pelo código do setor.
  - RP: em 2025 a coluna vem por extenso ("PRIMÁRIO DISCRICIONÁRIO") e chamada
    "Identificador Primário"; nos demais anos vem como código ("2") na coluna
    "Identificador Primário (Cod)". O texto é convertido para o código
    (ver RP_POR_EXTENSO); um texto desconhecido é mantido como veio e reportado.
  - NOMES DE COLUNA: "Esfera"/"Esfera (Cod)" e "Autor (Cod)"/"Ação (Cod)"
    trocam de nome entre anos. A leitura usa a lista de sinônimos em ALIAS.

============================================================================
REGRA 1 — Comando Militar de Área (C Mil A)
============================================================================
Mapeamento UF -> C Mil A. Caso especial de MINAS GERAIS: os municípios de
Uberlândia e Araguari pertencem ao CMP; o restante do estado ao CML.
Para registros com Autor (UF) = MG, o município é procurado (nesta ordem)
nas colunas "Localidade", "Subtítulo" e "Emenda (Justificativa)".

FALLBACK documentado: se não for possível identificar o município de um
registro de MG, o registro é atribuído ao CML, pois o CML cobre a quase
totalidade do território mineiro (todos os municípios exceto Uberlândia e
Araguari). O registro recebe a flag "cmilaFallback": true para transparência.

Autores sem UF (comissões, Autor (UF) = "NA") recebem "NÃO SE APLICA".

============================================================================
REGRA 2 — Inconsistências das emendas (aba "Inconsistências" do app)
============================================================================
São duas regras independentes; um mesmo registro pode acumular as duas.

--- Regra 2.A — Modalidade de Aplicação ---------------------------------
As emendas do Ministério da Defesa são executadas em Aplicação Direta,
portanto "Mod. Aplic. (Cod)" deve ser 90. Qualquer valor diferente de 90
(99 = "a definir", 91 = transferência a outro ente, etc.) é sinalizado como
inconsistência CONFIRMADA — é um dado objetivo, não interpretativo.

--- Regra 2.B — UO x Justificativa --------------------------------------
Cruza o texto de "Emenda (Justificativa)" (mais Ação e Subtítulo) com a UO
da emenda, para detectar emendas possivelmente destinadas a uma unidade
orçamentária que não corresponde à Força da OM descrita no texto.

  1. As UOs do MD são agrupadas em FAMÍLIAS por Força:
       MARINHA     : 52131 (Com. Marinha), 52931 (Fundo Naval),
                     52932 (Fundo Ens. Prof. Marítimo), 52133 (SECIRM)
       EXERCITO    : 52121 (Com. Exército), 52221 (IMBEL)
       AERONAUTICA : 52111 (Com. Aeronáutica), 52911 (Fundo Aeronáutico)
       (neutras)   : 52101 (Adm. Direta), 52902 (Fundo HFA) — órgãos
                     conjuntos do MD; só são sinalizadas quando o texto
                     aponta para UMA Força específica (o recurso deveria
                     então estar na UO daquela Força).

  2. A Força citada no texto é identificada por evidências, em ordem de
     prioridade:
       (i)  radical do CNPJ citado (evidência forte e determinística):
              00.394.502 -> MARINHA | 00.394.452 -> EXERCITO
              00.394.429 -> AERONAUTICA
       (ii) padrões de nome de OM (ex.: "BASE AÉREA", "CAPITANIA",
            "ESCOLA DE APRENDIZES MARINHEIROS", "BATALHÃO DE CAÇADORES").
     Quando há CNPJ no texto, ele PREVALECE sobre os padrões de nome.

  3. Se a Força citada difere da família da UO, o registro é sinalizado.

  4. REVISÃO QUALITATIVA (dicionário REVISAO_MANUAL): a varredura por
     palavras-chave produz falsos positivos conhecidos (p. ex. "batalhão",
     "infantaria" e "artilharia" em unidades de FUZILEIROS NAVAIS, que são
     da Marinha; PROFESP/Soldado Cidadão, conduzidos pelo próprio MD).
     Os casos revisados um a um recebem:
       - "confirmada" + diagnóstico redigido e UO sugerida; ou
       - "descartada" + motivo (não aparecem na aba, mas ficam contados).
     Casos automáticos ainda não revisados entram como "a verificar".

Cada inconsistência é gravada como um objeto estruturado
({tipo, gravidade, rotulo, descricao, evidencia, uoSugerida}), o que permite
à aba "Inconsistências" filtrar, agrupar e explicar cada achado.

ATENÇÃO: a Regra 2.B interpreta texto livre e é a parte mais sensível do
pipeline. Toda alteração deve ser validada contra a base real.
============================================================================
"""
import json
import os
import re
import shutil
import sys
import time
import unicodedata
import urllib.error
import urllib.request

import openpyxl

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
REPO_DADOS = "andresilvatorres85-bit/emendas.apresentadas.ploa"
ORGAO_COD = "52000"
SETOR_NOME = "DEFESA"  # recorte por nome: o código mudou de 08 (2023) para 13
ABA_PREFERIDA = "Todos os Setores"  # tem as colunas extras "UF" e "Localidade"
SAIDA = os.path.join(os.path.dirname(__file__), "..", "public", "dados.json")

# Sinônimos de coluna entre os anos (chave = nome canônico usado no código).
ALIAS = {
    "Identificador Primário (Cod)": ["Identificador Primário", "RP", "RP (Cod)"],
    "Esfera (Cod)": ["Esfera"],
    "Autor (Cod)": ["Ação (Cod)"],
    "Autor (Tipo)": ["Autor Tipo"],  # 2019 escreve sem parênteses
    "OM": ["OM (agregado)"],
    "Objeto": ["Objeto (agregado)"],
}

# Aba auxiliar do consolidado com o de-para (Ano|Emenda) -> OM/Objeto. Serve de
# reserva quando a linha do ano vem sem as colunas preenchidas.
ABA_OM_OBJETO = "Fonte_OM_Objeto"

# Rótulo para o "Autor (Tipo)" que a planilha não traz (71 linhas de 2020).
TIPO_AUTOR_DESCONHECIDO = "NÃO INFORMADO"

# "Identificador Primário" por extenso (2025) -> código de RP usado no app.
RP_POR_EXTENSO = {
    "FINANCEIRO": "0",
    "PRIMARIO OBRIGATORIO": "1",
    "PRIMARIO DISCRICIONARIO": "2",
    "PRIMARIO DISCRICIONARIO - PAC": "3",
    "DISCRICIONARIA DECORRENTE DE EMENDAS INDIVIDUAIS DE EXECUCAO OBRIGATORIA": "6",
    "DISCRICIONARIA DECORRENTE DE EMENDAS DE BANCADA IMPOSITIVA": "7",
    "DISCRICIONARIA DECORRENTE DE EMENDA DE RELATOR": "8",
    "DISCRICIONARIA DECORRENTE DE EMENDAS DE COMISSAO": "9",
}

ANO_RE = re.compile(r"(20\d{2})")

COLUNAS = [
    "Setor (Cod)", "Setor", "Emenda", "Emenda (Modalidade)", "Emenda (Tipo)",
    "Autor (Cod)", "Autor", "Autor (Tipo)", "Autor (UF)", "Partido",
    "Órgão (Cod)", "Órgão", "UO (Cod)", "UO", "Funcional", "Função",
    "Subfunção", "Programa", "Ação", "Subtítulo", "UF", "Localidade",
    "Esfera (Cod)", "GND (Cod)", "Mod. Aplic. (Cod)", "ID Uso (Cod)",
    "Identificador Primário (Cod)", "Fonte (Cod)", "Valor Solicitado",
    "Emenda (Justificativa)",
]

# ---------------------------------------------------------------------------
# REGRA 1 — C Mil A
# ---------------------------------------------------------------------------
UF_CMILA = {
    # CMA — Comando Militar da Amazônia
    "AC": "CMA", "AM": "CMA", "RO": "CMA", "RR": "CMA",
    # CMAO — Comando Militar da Amazônia Oriental
    "AP": "CMAO", "MA": "CMAO", "PA": "CMAO",
    # CMNE — Comando Militar do Nordeste
    "AL": "CMNE", "BA": "CMNE", "CE": "CMNE", "PB": "CMNE",
    "PE": "CMNE", "PI": "CMNE", "RN": "CMNE", "SE": "CMNE",
    # CMO — Comando Militar do Oeste
    "MT": "CMO", "MS": "CMO",
    # CMP — Comando Militar do Planalto (+ Uberlândia/Araguari-MG)
    "DF": "CMP", "GO": "CMP", "TO": "CMP",
    # CML — Comando Militar do Leste (+ MG exceto Uberlândia/Araguari)
    "ES": "CML", "RJ": "CML",
    # CMSE — Comando Militar do Sudeste
    "SP": "CMSE",
    # CMS — Comando Militar do Sul
    "PR": "CMS", "RS": "CMS", "SC": "CMS",
}
CMILA_NOMES = {
    "CMA": "Comando Militar da Amazônia",
    "CMAO": "Comando Militar da Amazônia Oriental",
    "CMNE": "Comando Militar do Nordeste",
    "CMO": "Comando Militar do Oeste",
    "CMP": "Comando Militar do Planalto",
    "CML": "Comando Militar do Leste",
    "CMSE": "Comando Militar do Sudeste",
    "CMS": "Comando Militar do Sul",
    "NÃO SE APLICA": "Sem UF de autor (comissões)",
}


def _sem_acento(txt):
    return unicodedata.normalize("NFKD", txt or "").encode("ascii", "ignore").decode()


def deduzir_cmila(registro):
    """Retorna (sigla_cmila, usou_fallback_mg)."""
    uf = (registro.get("Autor (UF)") or "").strip().upper()
    if uf == "MG":
        # Procura o município nas colunas Localidade, Subtítulo e Justificativa.
        texto = " | ".join(
            _sem_acento(str(registro.get(c) or "")).upper()
            for c in ("Localidade", "Subtítulo", "Emenda (Justificativa)")
        )
        if re.search(r"UBERLANDIA|ARAGUARI", texto):
            return "CMP", False
        # Município de MG identificado explicitamente e não é Uberlândia/Araguari?
        # Ex.: Localidade = 'SETE LAGOAS' (município) ou 'MINAS GERAIS' (estado).
        loc = _sem_acento(str(registro.get("Localidade") or "")).upper()
        municipio_identificado = loc not in ("", "MINAS GERAIS", "NACIONAL", "NA")
        # FALLBACK: sem município identificável -> CML (cobre MG exceto
        # Uberlândia e Araguari). Flag de fallback para transparência.
        return "CML", (not municipio_identificado)
    if uf in UF_CMILA:
        return UF_CMILA[uf], False
    return "NÃO SE APLICA", False


# ---------------------------------------------------------------------------
# REGRA 2 — Inconsistências (2.A Modalidade de Aplicação | 2.B UO x Justificativa)
# ---------------------------------------------------------------------------
MOD_APLIC_ESPERADA = "90"
MOD_APLIC_NOMES = {
    "90": "Aplicação Direta",
    "91": "Aplicação Direta decorrente de operação entre órgãos",
    "96": "Aplicação Direta decorrente de operação com consórcio público",
    "99": "A Definir",
    "30": "Transferência a Estados e ao DF",
    "31": "Transferência a Estados e ao DF — fundo a fundo",
    "32": "Execução Orçamentária Delegada a Estados e ao DF",
    "40": "Transferência a Municípios",
    "41": "Transferência a Municípios — fundo a fundo",
    "42": "Execução Orçamentária Delegada a Municípios",
    "50": "Transferência a Instituições Privadas sem Fins Lucrativos",
}

UO_FAMILIA = {
    "52131": "MARINHA", "52931": "MARINHA", "52932": "MARINHA", "52133": "MARINHA",
    "52232": "MARINHA",  # CCCPM
    "52233": "MARINHA",  # AMAZUL
    "52121": "EXERCITO", "52221": "EXERCITO", "52921": "EXERCITO",
    "52222": "EXERCITO",  # Fundação Osório
    "52111": "AERONAUTICA", "52911": "AERONAUTICA",
    "52211": "AERONAUTICA",  # Caixa de Financiamento Imobiliário da Aeronáutica
    # neutras (órgãos conjuntos do MD)
    "52101": None, "52902": None, "52901": None, "52903": None,
}
FAMILIA_LABEL = {"MARINHA": "Marinha", "EXERCITO": "Exército", "AERONAUTICA": "Aeronáutica"}
# Rótulo da Força usado no filtro "Órgão" do app.
FAMILIA_ORGAO = {
    "MARINHA": "MARINHA", "EXERCITO": "EXÉRCITO", "AERONAUTICA": "AERONÁUTICA",
    None: "MINISTÉRIO DA DEFESA",
}

# ---------------------------------------------------------------------------
# Classificação de UO ainda não catalogada
# ---------------------------------------------------------------------------
# Uma UO nova na planilha não pode virar "Ministério da Defesa" no silêncio —
# ela sumiria dentro dos órgãos conjuntos e distorceria o comparativo por Força.
# Três camadas, da mais forte para a mais fraca:
#   1. UO_FAMILIA — catálogo conferido à mão, sempre vence;
#   2. NOME da UO — "FUNDO NAVAL", "IMBEL", "FUNDO AERONÁUTICO" são inequívocos;
#   3. ESTRUTURA DO CÓDIGO — dentro do órgão 52000 o 4º dígito identifica o
#      Comando: 1=Aeronáutica, 2=Exército, 3=Marinha, 0=MD (órgão conjunto).
#      Vale para as 12 UOs conhecidas (52111/52911, 52121/52221/52921,
#      52131/52133/52232/52931/52932, 52101/52902).
# Toda UO que não estiver no catálogo é registrada em `uosNaoCatalogadas`
# (bloco `auditoria` do JSON) com a camada que a classificou.
NOME_FAMILIA = [
    (r"MARINHA|\bNAVAL\b|MARITIMO|RECURSOS DO MAR|FUZILEIROS|\bCCCPM\b|\bSECIRM\b"
     r"|\bAMAZUL\b|AMAZONIA AZUL", "MARINHA"),
    (r"EXERCITO|\bIMBEL\b|BELICO|\bOSORIO\b", "EXERCITO"),
    (r"AERONAUTICA|AEROESPACIAL|FORCA AEREA|\bAEREA\b", "AERONAUTICA"),
]
DIGITO_FAMILIA = {"0": None, "1": "AERONAUTICA", "2": "EXERCITO", "3": "MARINHA"}


def familia_da_uo(uo_cod, uo_nome, registro_desconhecidas=None):
    """Devolve a família (Força) de uma UO: MARINHA | EXERCITO | AERONAUTICA | None."""
    cod = str(uo_cod or "").strip()
    if cod in UO_FAMILIA:
        return UO_FAMILIA[cod]

    nome = _sem_acento(str(uo_nome or "")).upper()
    for padrao, familia in NOME_FAMILIA:
        if re.search(padrao, nome):
            if registro_desconhecidas is not None:
                registro_desconhecidas[cod] = {
                    "uo": str(uo_nome or ""), "familia": FAMILIA_ORGAO[familia], "criterio": "nome",
                }
            return familia

    if re.fullmatch(r"52\d{3}", cod):
        familia = DIGITO_FAMILIA.get(cod[3])
        if registro_desconhecidas is not None:
            registro_desconhecidas[cod] = {
                "uo": str(uo_nome or ""), "familia": FAMILIA_ORGAO[familia],
                "criterio": "4º dígito do código",
            }
        return familia

    if registro_desconhecidas is not None:
        registro_desconhecidas[cod] = {
            "uo": str(uo_nome or ""), "familia": FAMILIA_ORGAO[None], "criterio": "sem critério",
        }
    return None
FAMILIA_UO_SUGERIDA = {
    "MARINHA": "52131 (Comando da Marinha) ou 52931 (Fundo Naval)",
    "EXERCITO": "52121 (Comando do Exército)",
    "AERONAUTICA": "52111 (Comando da Aeronáutica) ou 52911 (Fundo Aeronáutico)",
}

CNPJ_FORCA = {"00394502": "MARINHA", "00394452": "EXERCITO", "00394429": "AERONAUTICA"}
CNPJ_RE = re.compile(r"(\d{2})[.\s]?(\d{3})[.\s]?(\d{3})\s*/\s*\d{4}\s*-?\s*\d{2}")

# Padrões de nome de OM por Força (texto sem acentos, maiúsculo).
# NOTA: termos genéricos que geram falso positivo entre Forças foram
# deliberadamente EXCLUÍDOS — "batalhão", "infantaria", "artilharia" e
# "companhia" existem tanto no Exército quanto no Corpo de Fuzileiros Navais
# (Marinha); "ala" aparece em textos comuns ("sala", "palavra"); "esquadrão"
# existe na Aeronáutica e na Cavalaria do Exército.
PADROES_OM = {
    "MARINHA": [
        r"MARINHA DO BRASIL", r"\bMARINHA\b", r"CAPITANIA DOS PORTOS",
        r"DELEGACIA FLUVIAL", r"FUZILEIROS NAVAIS", r"HOSPITAL NAVAL",
        r"\bSECIRM\b", r"AMAZONIA AZUL", r"\bSISGAAZ\b",
        r"APRENDIZES[ -]?MARINHEIROS", r"DISTRITO NAVAL", r"BASE NAVAL",
        r"\bNAVAL\b",
    ],
    "EXERCITO": [
        r"EXERCITO BRASILEIRO", r"\bEXERCITO\b", r"BATALHAO DE CACADORES",
        r"\bREGIAO MILITAR", r"COMANDO MILITAR D[AOE]", r"COLEGIO MILITAR",
        r"AGULHAS NEGRAS", r"PELOTAO ESPECIAL DE FRONTEIRA", r"\bSISFRON\b",
    ],
    "AERONAUTICA": [
        r"AERONAUTICA", r"\bFAB\b", r"FORCA AEREA", r"BASE AEREA",
        r"AERODROMO", r"\bCINDACTA\b", r"\bALA \d+\b", r"CADETES DO AR",
    ],
}

# ---------------------------------------------------------------------------
# REVISÃO QUALITATIVA (item 4 da Regra 2.B)
# ---------------------------------------------------------------------------
# Cada alerta automático da Regra 2.B foi lido caso a caso sobre a base do
# PLOA 2026. A chave é (nº da emenda, UO (Cod)) — estável entre cargas, ao
# contrário do nº da linha da planilha.
#   status "confirmada" -> divergência real; usa `descricao` e `uoSugerida`
#   status "descartada" -> falso positivo; não vai para a aba (fica no log)
# `gravidade`: "alta" = divergência clara | "media" = requer verificação
REVISAO_MANUAL = {
    ("2026", "41440013", "52101"): {
        "status": "confirmada", "gravidade": "alta",
        "descricao": "A UO é 52101 (Ministério da Defesa — Administração Direta), "
                     "mas a justificativa destina o recurso ao 28º Batalhão de Caçadores "
                     "do Exército Brasileiro, em Aracaju/SE.",
        "uoSugerida": "52121 (Comando do Exército)",
    },
    ("2026", "42510019", "52121"): {
        "status": "confirmada", "gravidade": "alta",
        "descricao": "A UO é 52121 (Comando do Exército), mas o objeto da emenda é a "
                     "Escola de Aprendizes-Marinheiros de Santa Catarina, organização "
                     "militar da Marinha do Brasil.",
        "uoSugerida": "52131 (Comando da Marinha) ou 52931 (Fundo Naval)",
    },
    ("2026", "44300001", "52121"): {
        "status": "confirmada", "gravidade": "alta",
        "descricao": "A UO é 52121 (Comando do Exército), mas a justificativa afirma "
                     "expressamente que a emenda \"visa atender a Marinha\" e descreve "
                     "material para o Corpo de Fuzileiros Navais e para um complexo naval.",
        "uoSugerida": "52131 (Comando da Marinha) ou 52931 (Fundo Naval)",
    },
    ("2026", "44460018", "52911"): {
        "status": "confirmada", "gravidade": "media",
        "descricao": "A UO é 52911 (Fundo Aeronáutico), em ação de funcionamento de "
                     "estabelecimentos de ensino profissional militares do MD, mas a "
                     "justificativa não identifica a OM beneficiada — apenas \"beneficiários "
                     "no estado de Minas Gerais\". Verificar se o destino não é o ensino "
                     "profissional marítimo (UO 52932) ou uma OM do Exército.",
        "uoSugerida": "verificar — possivelmente 52932 (Fundo de Ensino Profissional Marítimo)",
    },
    ("2026", "50270003", "52101"): {
        "status": "confirmada", "gravidade": "alta",
        "descricao": "A UO é 52101 (Ministério da Defesa — Administração Direta), mas a "
                     "justificativa informa o CNPJ 00.394.429 (COMANDO DA AERONÁUTICA) "
                     "como executor.",
        "uoSugerida": "52111 (Comando da Aeronáutica)",
    },
    ("2026", "60130005", "52101"): {
        "status": "confirmada", "gravidade": "alta",
        "descricao": "A UO é 52101 (Ministério da Defesa — Administração Direta), mas a "
                     "justificativa informa o CNPJ 00.394.429 (COMANDO DA AERONÁUTICA) "
                     "como executor.",
        "uoSugerida": "52111 (Comando da Aeronáutica)",
    },
    ("2026", "71020010", "52121"): {
        "status": "confirmada", "gravidade": "media",
        "descricao": "A UO é 52121 (Comando do Exército) para implantação/restauração de "
                     "aeródromo em Santa Rosa do Purus/AC. O município é sede de Pelotão "
                     "Especial de Fronteira do Exército, o que torna a UO plausível, mas a "
                     "competência sobre infraestrutura aeroportuária deve ser verificada "
                     "(Comando da Aeronáutica).",
        "uoSugerida": "verificar competência — 52111 (Comando da Aeronáutica)",
    },
    ("2026", "60020002", "52131"): {
        "status": "descartada",
        "motivo": "O Programa Fragatas Classe Tamandaré é conduzido pela Marinha; a "
                  "menção à indústria aeronáutica no texto é contextual.",
    },
}

# Falsos positivos genéricos já eliminados na própria varredura (documentado
# para memória do critério, ver comentário em PADROES_OM):
FALSOS_POSITIVOS_CONHECIDOS = [
    'Termos "batalhão", "infantaria", "artilharia" e "companhia" em unidades de '
    "Fuzileiros Navais (Marinha) — removidos dos padrões do Exército.",
    "PROFESP / Programa Soldado Cidadão — conduzidos pelo próprio Ministério da "
    "Defesa; a citação de uma Força é apenas o local da atividade.",
]


def forcas_citadas(texto_bruto):
    """Identifica as Forças das OMs citadas no texto.

    Retorna dict {familia: evidencia}. CNPJ prevalece: se houver ao menos um
    CNPJ de Força no texto, apenas os CNPJs são considerados (o nome de uma
    OM pode ser ambíguo entre Forças; o radical do CNPJ não é).
    """
    texto = _sem_acento(texto_bruto or "").upper().replace("_X000D_", " ")
    por_cnpj = {}
    for m in CNPJ_RE.finditer(texto):
        radical = "".join(m.groups())
        if radical in CNPJ_FORCA:
            por_cnpj.setdefault(CNPJ_FORCA[radical], f"CNPJ {m.group(0).strip()}")
    if por_cnpj:
        return por_cnpj
    por_nome = {}
    for familia, padroes in PADROES_OM.items():
        for padrao in padroes:
            m = re.search(padrao, texto)
            if m:
                por_nome.setdefault(familia, f'menção a "{m.group(0).strip()}"')
                break
    return por_nome


def detectar_inconsistencias(registro, descartadas):
    """Aplica as Regras 2.A e 2.B a um registro.

    Retorna a lista de inconsistências estruturadas. `descartadas` é uma lista
    acumuladora dos alertas revisados e descartados (só para o log/estatística).
    """
    achados = []
    ano = str(registro.get("__ano") or "").strip()
    emenda = str(registro.get("Emenda") or "").strip()
    uo_cod = str(registro.get("UO (Cod)") or "").strip()
    uo_nome = str(registro.get("UO") or "").strip()

    # --- Regra 2.A — Mod. Aplic. deve ser 90 --------------------------------
    mod = str(registro.get("Mod. Aplic. (Cod)") or "").strip()
    if mod and mod != MOD_APLIC_ESPERADA:
        nome_mod = MOD_APLIC_NOMES.get(mod, "modalidade não prevista")
        achados.append({
            "tipo": "modalidade",
            "gravidade": "alta",
            "rotulo": f"Mod. Aplic. {mod}",
            "descricao": (
                f"Modalidade de Aplicação {mod} ({nome_mod}). As emendas do Ministério "
                f"da Defesa são executadas em Aplicação Direta, portanto o código "
                f"esperado é 90."
            ),
            "evidencia": f"Mod. Aplic. (Cod) = {mod}",
            "uoSugerida": "",
        })

    # --- Regra 2.B — UO x Justificativa -------------------------------------
    familia_uo = familia_da_uo(uo_cod, uo_nome)
    texto = " ".join(str(registro.get(c) or "") for c in
                     ("Ação", "Subtítulo", "Emenda (Justificativa)"))
    citadas = forcas_citadas(texto)
    # A chave inclui o ano: o número da emenda repete entre exercícios (é o
    # código do autor + sequencial), então sem o ano uma revisão de 2026
    # poderia ser aplicada por engano a uma emenda de outro ano.
    revisao = REVISAO_MANUAL.get((ano, emenda, uo_cod))

    divergentes = {f: e for f, e in citadas.items() if f != familia_uo}
    if familia_uo is None:
        # UO neutra (órgão conjunto): só é indício quando o texto aponta para
        # UMA única Força — aí o recurso deveria estar na UO daquela Força.
        divergentes = citadas if len(citadas) == 1 else {}

    if revisao and revisao["status"] == "descartada":
        if divergentes:
            descartadas.append({"ano": ano, "emenda": emenda, "uoCod": uo_cod,
                                "motivo": revisao["motivo"]})
        divergentes = {}

    if revisao and revisao["status"] == "confirmada":
        # Casos confirmados na revisão qualitativa entram mesmo quando a
        # varredura automática não encontra palavra-chave (a divergência pode
        # estar na ausência de identificação da OM — ver emenda 44460018).
        forca = (sorted(divergentes)[0] if divergentes else None)
        rotulo_forca = FAMILIA_LABEL[forca] if forca else "verificar destino"
        achados.append({
            "tipo": "uo_justificativa",
            "gravidade": revisao["gravidade"],
            "rotulo": f"UO × justificativa — {rotulo_forca}",
            "descricao": revisao["descricao"],
            "evidencia": divergentes.get(forca, "análise do objeto descrito na justificativa"),
            "uoSugerida": revisao["uoSugerida"],
            "forcaUO": FAMILIA_LABEL.get(familia_uo, "Ministério da Defesa"),
            "forcaCitada": rotulo_forca,
            "revisado": True,
        })
    elif divergentes:
        # Alerta automático, ainda sem revisão qualitativa registrada.
        forca, evidencia = sorted(divergentes.items())[0]
        rotulo_forca = FAMILIA_LABEL[forca]
        origem = (f'a UO da emenda é "{uo_cod} — {uo_nome}"'
                  + (f" ({FAMILIA_LABEL[familia_uo]})" if familia_uo else
                     " (órgão conjunto do MD)"))
        achados.append({
            "tipo": "uo_justificativa",
            "gravidade": "media",
            "rotulo": f"UO × justificativa — {rotulo_forca}",
            "descricao": (
                f"O texto da emenda cita organização militar da {rotulo_forca} "
                f"({evidencia}), mas {origem}. Alerta automático ainda não revisado "
                f"manualmente — verificar."
            ),
            "evidencia": evidencia,
            "uoSugerida": FAMILIA_UO_SUGERIDA[forca],
            "forcaUO": FAMILIA_LABEL.get(familia_uo, "Ministério da Defesa"),
            "forcaCitada": rotulo_forca,
            "revisado": False,
        })

    return achados


# ---------------------------------------------------------------------------
# Leitura dos .xlsx
# ---------------------------------------------------------------------------
# Token do GitHub (fornecido automaticamente pelo Actions como GITHUB_TOKEN).
# Autenticar eleva o limite de requisições da API de 60/h (anônimo, por IP —
# facilmente estourado nos runners compartilhados) para 5.000/h.
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN") or ""


def _abrir(url, tentativas=4):
    """urlopen com User-Agent, token (quando houver) e retry com backoff em
    caso de 403 (rate limit) ou erros transitórios."""
    req = urllib.request.Request(url, headers={"User-Agent": "emendas-defesa-app-build"})
    host = url.split("/")[2] if "//" in url else ""
    if GITHUB_TOKEN and host.endswith(("github.com", "githubusercontent.com")):
        req.add_header("Authorization", f"Bearer {GITHUB_TOKEN}")
        req.add_header("X-GitHub-Api-Version", "2022-11-28")
    ultimo_erro = None
    for i in range(tentativas):
        try:
            return urllib.request.urlopen(req)
        except urllib.error.HTTPError as e:
            ultimo_erro = e
            # 403/429 = rate limit/abuso; 5xx = transitório. Aguarda e tenta de novo.
            if e.code in (403, 429) or 500 <= e.code < 600:
                espera = 2 ** i * 5
                print(f"  {e.code} em {url} — nova tentativa em {espera}s ({i+1}/{tentativas})")
                time.sleep(espera)
                continue
            raise
    raise ultimo_erro


def listar_xlsx_github():
    url = f"https://api.github.com/repos/{REPO_DADOS}/contents/"
    with _abrir(url) as r:
        itens = json.load(r)
    return [i["download_url"] for i in itens if i["name"].lower().endswith(".xlsx")]


def baixar(url, destino):
    print(f"Baixando {url}")
    with _abrir(url) as r, open(destino, "wb") as f:
        shutil.copyfileobj(r, f)
    return destino


def _resolver_alias(registro):
    """Preenche os nomes canônicos de coluna a partir dos sinônimos do ano."""
    for canonico, sinonimos in ALIAS.items():
        if registro.get(canonico) not in (None, ""):
            continue
        for alt in sinonimos:
            if registro.get(alt) not in (None, ""):
                registro[canonico] = registro[alt]
                break
    return registro


def _normalizar_rp(registro, desconhecidos):
    """RP sempre como código. Em 2025 a planilha traz o texto por extenso."""
    bruto = str(registro.get("Identificador Primário (Cod)") or "").strip()
    if not bruto or re.fullmatch(r"\d+", bruto):
        return
    chave = re.sub(r"\s+", " ", _sem_acento(bruto).upper()).strip()
    cod = RP_POR_EXTENSO.get(chave)
    if cod:
        registro["Identificador Primário (Cod)"] = cod
    else:
        desconhecidos.add(bruto)


def _ler_aba(ws, ano, desconhecidos):
    """Devolve os registros do Órgão 52000 / setor DEFESA de uma aba."""
    linhas = ws.iter_rows(values_only=True)
    cabecalho = [str(c).strip() if c is not None else "" for c in next(linhas)]
    registros = []
    for linha in linhas:
        d = dict(zip(cabecalho, linha))
        # Recorte de escopo. O setor é conferido pelo NOME porque o código
        # mudou de 08 (2023) para 13 (2024 em diante) — ver REGRA 0.
        if str(d.get("Órgão (Cod)") or "").strip() != ORGAO_COD:
            continue
        if _sem_acento(str(d.get("Setor") or "")).strip().upper() != SETOR_NOME:
            continue
        _resolver_alias(d)
        _normalizar_rp(d, desconhecidos)
        d["__ano"] = ano
        registros.append(d)
    return registros


def _ler_om_objeto(wb):
    """De-para (ano, emenda) -> (OM, Objeto) da aba auxiliar do consolidado.

    A coluna OM/Objeto já vem preenchida em cada aba de ano; esta aba é a
    reserva, usada quando a linha do ano está vazia. Só entra quem tem OM ou
    objeto de fato — a aba tem colunas de trabalho (chaves, agregados) que não
    interessam aqui.
    """
    if ABA_OM_OBJETO not in wb.sheetnames:
        return {}
    ws = wb[ABA_OM_OBJETO]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = [str(c).strip() if c is not None else "" for c in next(linhas)]
    mapa = {}
    for linha in linhas:
        d = dict(zip(cabecalho, linha))
        _resolver_alias(d)
        ano = str(d.get("Ano") or "").strip()
        emenda = str(d.get("Emenda") or "").strip()
        om = str(d.get("OM") or "").strip()
        objeto = str(d.get("Objeto") or "").strip()
        if ano and emenda and (om or objeto):
            mapa.setdefault((ano, emenda), (om, objeto))
    return mapa


def ler_blocos(caminho_xlsx, desconhecidos, om_objeto=None):
    """Lê um .xlsx e devolve [(ano, origem, registros)].

    `origem` é "exercicio" (arquivo de um único ano) ou "historico" (arquivo
    consolidado com uma aba por ano) — usado para desempatar anos repetidos.
    Abas que não são um ano (p. ex. `Fonte_OM_Objeto`) não viram bloco: elas
    são lidas à parte, como fonte auxiliar.
    """
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    if om_objeto is not None:
        om_objeto.update(_ler_om_objeto(wb))
    abas_ano = [n for n in wb.sheetnames if re.fullmatch(r"\s*20\d{2}\s*", str(n))]
    if abas_ano:
        blocos = []
        for nome in abas_ano:
            ano = str(nome).strip()
            registros = _ler_aba(wb[nome], ano, desconhecidos)
            print(f"  aba {ano}: {len(registros)} registros no escopo")
            blocos.append((ano, "historico", registros))
        return blocos

    # Arquivo de um único exercício: o ano vem do nome do arquivo.
    m = ANO_RE.search(os.path.basename(caminho_xlsx))
    if not m:
        print(f"  AVISO: não foi possível identificar o ano de {caminho_xlsx}; arquivo ignorado")
        return []
    ano = m.group(1)
    # Prefere a aba consolidada (tem "UF" e "Localidade"); senão, "Setor 13".
    if ABA_PREFERIDA in wb.sheetnames:
        ws = wb[ABA_PREFERIDA]
    elif "Setor 13" in wb.sheetnames:
        ws = wb["Setor 13"]
    else:
        ws = wb[wb.sheetnames[0]]
    registros = _ler_aba(ws, ano, desconhecidos)
    print(f"  aba utilizada: {ws.title} — ano {ano}, {len(registros)} registros no escopo")
    return [(ano, "exercicio", registros)]


def uniformizar_entre_anos(por_ano):
    """Alinha valores que a planilha escreve de formas diferentes entre anos.

    Duas correções, ambas visíveis direto nos filtros do app:

    1. NOME DA UO. O mesmo código aparece como "COMANDO DO EXÉRCITO" (2021+) e
       "COMANDO DO EXÉRCITO - ADMINISTRAÇÃO DIRETA" (2019/2020). Sem alinhar, o
       filtro "UO" lista a mesma unidade duas vezes e divide os totais. Vence o
       nome do ANO MAIS RECENTE em que o código aparece — assim o rótulo segue o
       que a planilha usa hoje, sem mapa escrito à mão.

    2. TIPO DE AUTOR AUSENTE. Em 2020, 71 linhas vêm sem "Autor (Tipo)", embora
       tragam o parlamentar e a UF. O tipo é preenchido a partir de outra linha
       do MESMO autor (em qualquer ano) — é o mesmo parlamentar, e sem isso ele
       fica de fora do ranking e da rosca de impositivas.
    """
    anos_desc = sorted(por_ano, reverse=True)

    nome_uo = {}
    tipo_autor = {}
    for ano in anos_desc:
        for r in por_ano[ano][1]:
            cod = str(r.get("UO (Cod)") or "").strip()
            nome = str(r.get("UO") or "").strip()
            if cod and nome:
                nome_uo.setdefault(cod, nome)
            autor = str(r.get("Autor") or "").strip().upper()
            tipo = str(r.get("Autor (Tipo)") or "").strip()
            if autor and tipo:
                tipo_autor.setdefault(autor, tipo)

    renomeadas, tipos_preenchidos, sem_tipo = set(), 0, 0
    for ano in por_ano:
        for r in por_ano[ano][1]:
            cod = str(r.get("UO (Cod)") or "").strip()
            canonico = nome_uo.get(cod)
            if canonico and str(r.get("UO") or "").strip() != canonico:
                renomeadas.add((cod, str(r.get("UO") or "").strip(), canonico))
                r["UO"] = canonico
            if not str(r.get("Autor (Tipo)") or "").strip():
                tipo = tipo_autor.get(str(r.get("Autor") or "").strip().upper())
                if tipo:
                    r["Autor (Tipo)"] = tipo
                    tipos_preenchidos += 1
                elif str(r.get("Autor") or "").strip():
                    # Sem evidência de qual Casa: rotular explicitamente é melhor
                    # que deixar vazio (vira uma opção em branco no filtro) e do
                    # que adivinhar. O código do autor NÃO serve de pista — ele é
                    # reaproveitado entre legislaturas (5 códigos aparecem ora
                    # como Deputado, ora como Senador).
                    r["Autor (Tipo)"] = TIPO_AUTOR_DESCONHECIDO
                    sem_tipo += 1

    if renomeadas:
        print("\nNome de UO alinhado ao do ano mais recente:")
        for cod, antigo, novo in sorted(renomeadas):
            print(f"  {cod}: {antigo!r} -> {novo!r}")
    if tipos_preenchidos:
        print(f"\nAutor (Tipo) ausente preenchido pelo mesmo autor em outro ano: "
              f"{tipos_preenchidos} linha(s)")
    if sem_tipo:
        print(f"\nAutor (Tipo) ausente e sem evidência em outro ano: {sem_tipo} "
              f"linha(s) rotuladas como {TIPO_AUTOR_DESCONHECIDO!r}")


def normalizar(registro, idx, descartadas, uos_nao_catalogadas=None, om_objeto=None):
    """Converte um registro bruto no formato compacto consumido pelo app."""
    def s(col):
        v = registro.get(col)
        return "" if v is None else str(v).strip()

    valor = registro.get("Valor Solicitado")
    if isinstance(valor, str):
        valor = float(re.sub(r"[^\d,.-]", "", valor).replace(".", "").replace(",", ".") or 0)
    valor = float(valor or 0)

    cmila, fallback_mg = deduzir_cmila(registro)
    inconsistencias = detectar_inconsistencias(registro, descartadas)
    justificativa = s("Emenda (Justificativa)").replace("_x000D_", "\n").strip()

    # A Força vem calculada do pipeline: é aqui que uma UO nova é classificada,
    # em um lugar só, e o app apenas lê o campo.
    orgao = FAMILIA_ORGAO[familia_da_uo(s("UO (Cod)"), s("UO"), uos_nao_catalogadas)]

    out = {
        "id": idx,
        "ano": str(registro.get("__ano") or ""),
        "emenda": s("Emenda"),
        "orgao": orgao,
        "modalidade": s("Emenda (Modalidade)"),
        "tipo": s("Emenda (Tipo)"),
        "autor": s("Autor"),
        "autorTipo": s("Autor (Tipo)"),
        "autorUF": s("Autor (UF)"),
        "partido": s("Partido"),
        "uoCod": s("UO (Cod)"),
        "uo": s("UO"),
        "funcional": s("Funcional"),
        "acao": s("Ação"),
        "subtitulo": s("Subtítulo"),
        "localidade": s("Localidade") or s("Subtítulo"),
        "gnd": s("GND (Cod)"),
        "modAplic": s("Mod. Aplic. (Cod)"),
        "rp": s("Identificador Primário (Cod)"),
        "valor": valor,
        "justificativa": justificativa,
        "cmila": cmila,
    }

    # OM beneficiada e objeto da emenda. Vêm das colunas da própria aba do ano;
    # a aba auxiliar do consolidado é a reserva. Só existem onde alguém já fez a
    # identificação — hoje, quase toda em emendas do Exército.
    om = s("OM")
    objeto = s("Objeto")
    if not (om or objeto) and om_objeto:
        om, objeto = om_objeto.get((out["ano"], out["emenda"]), ("", ""))
    if om:
        out["om"] = om.replace("\xa0", " ").strip()
    if objeto:
        out["objeto"] = objeto.replace("_x000D_", " ").replace("\xa0", " ").strip()
    if fallback_mg:
        out["cmilaFallback"] = True
    if inconsistencias:
        out["inconsistencias"] = inconsistencias
    return out


# ---------------------------------------------------------------------------
# REGRA 3 — PLOA: despesas por fase de elaboração
# ---------------------------------------------------------------------------
# Fonte: `PLOA_Despesas_Elaboracao.xlsx`, uma aba por exercício, com o valor de
# cada dotação em cada fase da tramitação do projeto de lei orçamentária.
# Escopo: TODO o órgão 52000 — aqui não há recorte por setor, porque a pergunta
# da aba PLOA é "quanto foi para o Ministério da Defesa", e parte das UO do
# órgão (a ALADA, por exemplo) responde pelo setor de Ciência & Tecnologia.
#
# 3.A — Fase ausente fica EM BRANCO (não herda a fase anterior).
#   A planilha não preenche a mesma sequência de colunas em todos os anos: no
#   início do rito o exercício corrente só tem o PL, 2022 pára no Ciclo Plenário
#   (Autógrafo ausente) e 2023 pula o Ciclo Plenário. A ausência é do ANO, não da
#   linha — um zero numa linha isolada é um zero de verdade (dotação criada só no
#   Ciclo Geral começa com PL = 0, e isso é informação). Por isso a detecção é
#   feita por COLUNA: fase cuja coluna inexiste, ou soma zero no ano inteiro, é
#   considerada ausente e recebe `null` (branco). Herdar o valor da fase anterior
#   afirmaria um dado que a planilha ainda não tem; deixar em branco é honesto e é
#   o que a análise "desde o início do rito" pede — a leitura primária dos painéis
#   passou a ser o PL, sempre presente, e o autógrafo entra só quando existe. O
#   JSON guarda os valores crus (`fases`) e os efetivos com branco (`fasesEf`),
#   mais a lista de fases ausentes por ano (`fasesVazias`) — assim a tela sabe
#   quais fases exibir em branco.
#
# 3.B — Anos com conteúdo idêntico são sinalizados, não removidos.
#   Hoje a aba 2025 é cópia bit a bit da 2024 (mesmas 669 linhas, mesmos cinco
#   valores). Descartar em silêncio esconderia o defeito; somar como se fossem
#   exercícios distintos mentiria. O pipeline registra o par em
#   `auditoria.anosDuplicados` e o app mostra o aviso na própria aba.
FASES = [
    ("pl", "PL"),
    ("setorial", "Ciclo Setorial"),
    ("geral", "Ciclo Geral"),
    ("plenario", "Ciclo Plenário"),
    ("autografo", "Autógrafo"),
]
FASE_IDS = [f[0] for f in FASES]
FASE_ROTULOS = {f[0]: f[1] for f in FASES}

# Sinônimos de coluna do arquivo do PLOA. Escopo próprio: no arquivo das
# emendas "Subtítulo" é o localizador e não pode ser confundido com a
# subfunção — e é exatamente essa a diferença entre as abas aqui (2024 traz
# "Subtítulo" e não traz "Subfunção").
ALIAS_PLOA = {
    "Órgão (Cod)": ["Órgao (Cod)"],  # a aba do PLOA escreve sem o til
    "Identificador Primário (Cod)": ["Identificador Primário", "RP", "RP (Cod)"],
}

GND_NOMES = {
    "1": "Pessoal e encargos sociais",
    "2": "Juros e encargos da dívida",
    "3": "Outras despesas correntes",
    "4": "Investimentos",
    "5": "Inversões financeiras",
    "6": "Amortização da dívida",
    "9": "Reserva de contingência",
}


def _eh_planilha_ploa(caminho_xlsx):
    """Distingue o arquivo do PLOA do arquivo das emendas pelo cabeçalho.

    O reconhecimento é pelo FORMATO, não pelo nome do arquivo: renomear a
    planilha na origem não quebra o pipeline, e um arquivo novo com o mesmo
    formato é absorvido sozinho. A assinatura são as colunas de fase, que só
    existem no arquivo de elaboração.
    """
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    try:
        for nome in wb.sheetnames:
            linhas = wb[nome].iter_rows(values_only=True)
            try:
                cabecalho = {str(c).strip() for c in next(linhas) if c is not None}
            except StopIteration:
                continue
            if {"PL", "Ciclo Geral", "Autógrafo"} <= cabecalho:
                return True
        return False
    finally:
        wb.close()


def _segmento_funcional(funcional, i):
    """Segmento `i` do código funcional FF.SSS.PPPP.AAAA.LLLL.

    Serve de reserva para as colunas que faltam em algum ano — 2024 não traz
    "Subfunção", e o código funcional traz. Derivar do funcional é melhor que
    um mapa por ano: vale para qualquer aba nova.
    """
    partes = str(funcional or "").split(".")
    return partes[i].strip() if len(partes) > i else ""


def _num(v):
    """Valor monetário da célula. Texto com máscara ou vazio vira 0."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    limpo = str(v).strip().replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return 0.0


def ler_ploa(caminho_xlsx, uos_nao_catalogadas=None):
    """Lê o arquivo de elaboração e devolve (registros, fases_vazias_por_ano)."""
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    registros, fases_vazias = [], {}
    abas_ano = [n for n in wb.sheetnames if re.fullmatch(r"\s*20\d{2}\s*", str(n))]
    for nome in abas_ano:
        ano = str(nome).strip()
        ws = wb[nome]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [str(c).strip() if c is not None else "" for c in next(linhas)]
        do_ano = []
        for linha in linhas:
            d = dict(zip(cabecalho, linha))
            for canonico, sinonimos in ALIAS_PLOA.items():
                if not d.get(canonico):
                    for s in sinonimos:
                        if d.get(s) not in (None, ""):
                            d[canonico] = d[s]
                            break
            if str(d.get("Órgão (Cod)") or "").strip() != ORGAO_COD:
                continue
            funcional = str(d.get("Funcional") or "").strip()
            uo_cod = str(d.get("UO (Cod)") or "").strip()
            uo_nome = str(d.get("UO") or "").strip()
            familia = familia_da_uo(uo_cod, uo_nome, uos_nao_catalogadas)
            gnd = str(d.get("GND (Cod)") or "").strip()
            do_ano.append({
                "ano": ano,
                "setorCod": str(d.get("Setor (Cod)") or "").strip(),
                "setor": str(d.get("Setor") or "").strip(),
                "uoCod": uo_cod,
                "uo": uo_nome,
                "orgao": FAMILIA_ORGAO[familia],
                "uf": str(d.get("UF") or "").strip(),
                "funcional": funcional,
                "funcao": str(d.get("Função") or "").strip(),
                "funcaoCod": (str(d.get("Função (Cod)") or "").strip()
                              or _segmento_funcional(funcional, 0)),
                "subfuncao": str(d.get("Subfunção") or "").strip(),
                "subfuncaoCod": (str(d.get("Subfunção (Cod)") or "").strip()
                                 or _segmento_funcional(funcional, 1)),
                "programa": str(d.get("Programa") or "").strip(),
                "programaCod": (str(d.get("Programa (Cod)") or "").strip()
                                or _segmento_funcional(funcional, 2)),
                "acao": str(d.get("Ação") or "").strip(),
                "acaoCod": (str(d.get("Ação (Cod)") or "").strip()
                            or _segmento_funcional(funcional, 3)),
                "subtituloCod": (str(d.get("Subtítulo (Cod)") or "").strip()
                                 or _segmento_funcional(funcional, 4)),
                "gnd": gnd,
                "gndNome": GND_NOMES.get(gnd, ""),
                "rp": str(d.get("Identificador Primário (Cod)") or "").strip(),
                "modAplic": str(d.get("Mod. Aplic. (Cod)") or "").strip(),
                "fonte": str(d.get("Fonte (Cod)") or "").strip(),
                "fases": [_num(d.get(rot)) for _, rot in FASES],
            })

        # REGRA 3.A — a ausência é da COLUNA no ANO, nunca da linha.
        presentes = [rot in cabecalho for _, rot in FASES]
        vazias = [
            fid for i, (fid, _) in enumerate(FASES)
            if not presentes[i] or not any(r["fases"][i] for r in do_ano)
        ]
        fases_vazias[ano] = vazias
        indices_vazios = {FASE_IDS.index(f) for f in vazias}
        for r in do_ano:
            efetivas = []
            for i in range(len(FASES)):
                bruto = r["fases"][i]
                if i in indices_vazios and i > 0:
                    # REGRA 3.A — fase ausente fica em branco (não herda). O PL
                    # (i == 0) nunca é apagado: é o valor de partida do rito.
                    efetivas.append(None)
                else:
                    efetivas.append(bruto)
            r["fasesEf"] = efetivas
        print(f"  PLOA aba {ano}: {len(do_ano)} linhas do órgão {ORGAO_COD}"
              + (f" | fase(s) sem valor na planilha, exibidas em branco: "
                 f"{', '.join(FASE_ROTULOS[f] for f in vazias)}" if vazias else ""))
        registros.extend(do_ano)
    wb.close()
    return registros, fases_vazias


def anos_duplicados_ploa(registros):
    """REGRA 3.B — anos cujo conteúdo é idêntico ao de outro ano."""
    assinatura = {}
    for r in registros:
        chave = (r["uoCod"], r["funcional"], r["gnd"], r["rp"], r["fonte"],
                 r["modAplic"], tuple(r["fases"]))
        assinatura.setdefault(r["ano"], []).append(chave)
    resumo, pares = {a: tuple(sorted(v)) for a, v in assinatura.items()}, []
    anos = sorted(resumo)
    for i, a in enumerate(anos):
        for b in anos[i + 1:]:
            if resumo[a] == resumo[b]:
                pares.append({"ano": b, "igualA": a, "linhas": len(resumo[a])})
    return pares


def main():
    if len(sys.argv) > 1:
        pasta = sys.argv[1]
        arquivos = [
            os.path.join(pasta, f) for f in sorted(os.listdir(pasta))
            if f.lower().endswith(".xlsx") and not f.startswith("~$")
        ]
    else:
        os.makedirs("/tmp/xlsx_dados", exist_ok=True)
        arquivos = [
            baixar(u, os.path.join("/tmp/xlsx_dados", os.path.basename(u)))
            for u in listar_xlsx_github()
        ]
    if not arquivos:
        sys.exit("Nenhum arquivo .xlsx encontrado.")

    # O repositório de dados guarda duas planilhas de naturezas diferentes: as
    # emendas apresentadas e as despesas por fase de elaboração do PLOA. A
    # triagem é pelo cabeçalho (ver `_eh_planilha_ploa`) — não pelo nome —,
    # então uma planilha renomeada na origem continua caindo no lugar certo.
    arquivos_ploa = [a for a in arquivos if _eh_planilha_ploa(a)]
    arquivos = [a for a in arquivos if a not in arquivos_ploa]
    if not arquivos:
        sys.exit("Nenhuma planilha de emendas encontrada (só arquivos de PLOA).")

    # 1) Lê tudo e agrupa por ano. Ver REGRA 0 para a precedência entre o
    #    arquivo do exercício e a aba do consolidado.
    rp_desconhecidos = set()
    om_objeto = {}
    por_ano = {}
    for arq in arquivos:
        print(f"Lendo {arq}")
        for ano, origem, brutos in ler_blocos(arq, rp_desconhecidos, om_objeto):
            atual = por_ano.get(ano)
            if atual and atual[0] == "exercicio" and origem == "historico":
                print(f"  ano {ano}: aba do consolidado ignorada "
                      f"(já veio do arquivo do exercício)")
                continue
            if atual and origem == "exercicio" and atual[0] == "historico":
                print(f"  ano {ano}: arquivo do exercício substitui a aba do consolidado")
            elif atual:
                print(f"  ano {ano}: {len(brutos)} registros somados aos {len(atual[1])} já lidos")
                por_ano[ano] = (origem, atual[1] + brutos)
                continue
            por_ano[ano] = (origem, brutos)

    if not por_ano:
        sys.exit("Nenhum registro no escopo (Órgão 52000 / setor DEFESA).")
    if rp_desconhecidos:
        print("\nAVISO: RP por extenso não reconhecido (mantido como veio): "
              + "; ".join(sorted(rp_desconhecidos)))

    # Alinha entre anos o que a planilha escreve de formas diferentes, ANTES de
    # normalizar — o resultado vai direto para os filtros do app.
    uniformizar_entre_anos(por_ano)

    # 2) Normaliza ano a ano. A deduplicação é POR ANO: o mesmo número de
    #    emenda existe em exercícios diferentes e são emendas diferentes.
    registros, descartadas, uos_nao_catalogadas = [], [], {}
    for ano in sorted(por_ano):
        vistos = set()
        for r in por_ano[ano][1]:
            chave = tuple(str(r.get(c) or "") for c in COLUNAS if c in r)
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append(
                normalizar(r, len(registros), descartadas, uos_nao_catalogadas, om_objeto))

    anos = sorted({r["ano"] for r in registros})
    print("\nPor ano:")
    for ano in anos:
        do_ano = [r for r in registros if r["ano"] == ano]
        print(f"  {ano} ({por_ano[ano][0]}): {len(do_ano)} registros | "
              f"{len(set(r['emenda'] for r in do_ano))} emendas | "
              f"R$ {sum(r['valor'] for r in do_ano):,.2f}")

    if uos_nao_catalogadas:
        print("\nUO fora do catálogo (classificadas automaticamente — conferir):")
        for cod, info in sorted(uos_nao_catalogadas.items()):
            print(f"  {cod} {info['uo']} -> {info['familia']} (por {info['criterio']})")

    # 3) PLOA — despesas por fase de elaboração (REGRA 3). Independente das
    #    emendas: outro arquivo, outro escopo (o órgão inteiro) e outra chave.
    ploa_registros, ploa_fases_vazias, ploa_uos = [], {}, {}
    for arq in arquivos_ploa:
        print(f"\nLendo PLOA {arq}")
        regs, vazias = ler_ploa(arq, ploa_uos)
        ploa_registros.extend(regs)
        ploa_fases_vazias.update(vazias)
    ploa_anos = sorted({r["ano"] for r in ploa_registros})
    ploa_duplicados = anos_duplicados_ploa(ploa_registros)
    # Enxuga o JSON: 3.300 dotações × 25 campos pesam num app que também carrega
    # 4.600 registros de emenda. Campo vazio não vai, `gndNome` sai (há a tabela
    # `gndNomes` no cabeçalho do bloco) e `fases` só é emitido quando difere de
    # `fasesEf` — ou seja, apenas nos anos em que alguma fase ficou em branco.
    for r in ploa_registros:
        r.pop("gndNome", None)
        if r["fases"] == r["fasesEf"]:
            r.pop("fases")
        for chave in [k for k, v in r.items() if v == ""]:
            del r[chave]
    if ploa_registros:
        print("\nPLOA por ano (PL e autógrafo; autógrafo em branco = ainda não na planilha):")
        for ano in ploa_anos:
            do_ano = [r for r in ploa_registros if r["ano"] == ano]
            # fasesEf pode conter None (fase em branco): soma protegida.
            final = sum((r["fasesEf"][-1] or 0) for r in do_ano)
            pl = sum((r["fasesEf"][0] or 0) for r in do_ano)
            sem_aut = "autografo" in ploa_fases_vazias.get(ano, [])
            print(f"  {ano}: {len(do_ano)} linhas | PL R$ {pl:,.2f} | "
                  + ("autógrafo em branco" if sem_aut else f"autógrafo R$ {final:,.2f}"))
    for d in ploa_duplicados:
        print(f"  AVISO: a aba {d['ano']} é idêntica à aba {d['igualA']} "
              f"({d['linhas']} linhas iguais) — sinalizado no app, não removido")

    n_incons = sum(1 for r in registros if r.get("inconsistencias"))
    n_mod = sum(1 for r in registros
                if any(i["tipo"] == "modalidade" for i in r.get("inconsistencias", [])))
    n_uo = sum(1 for r in registros
               if any(i["tipo"] == "uo_justificativa" for i in r.get("inconsistencias", [])))
    saida = {
        "geradoEm": __import__("datetime").datetime.utcnow().isoformat() + "Z",
        "fonte": f"github.com/{REPO_DADOS}",
        "escopo": {"orgaoCod": ORGAO_COD, "orgao": "MINISTÉRIO DA DEFESA", "setor": SETOR_NOME},
        "anos": anos,
        "anoCorrente": anos[-1] if anos else "",
        "origemPorAno": {a: por_ano[a][0] for a in anos},
        "cmilaNomes": CMILA_NOMES,
        "auditoria": {
            "modAplicEsperada": MOD_APLIC_ESPERADA,
            "revisadosDescartados": len(descartadas),
            "uosNaoCatalogadas": uos_nao_catalogadas,
            "falsosPositivosConhecidos": FALSOS_POSITIVOS_CONHECIDOS,
            "anosDuplicadosPLOA": ploa_duplicados,
            "uosNaoCatalogadasPLOA": ploa_uos,
        },
        "registros": registros,
        # Bloco da aba PLOA. Vive separado de `registros` porque é outra base:
        # outro arquivo, outro escopo (órgão inteiro, todos os setores) e outra
        # unidade de análise (a dotação, não a emenda).
        "ploa": {
            "anos": ploa_anos,
            "anoCorrente": ploa_anos[-1] if ploa_anos else "",
            "fases": [{"id": i, "rotulo": r} for i, r in FASES],
            "fasesVazias": ploa_fases_vazias,
            "anosDuplicados": ploa_duplicados,
            "gndNomes": GND_NOMES,
            "registros": ploa_registros,
        },
    }
    os.makedirs(os.path.dirname(os.path.abspath(SAIDA)), exist_ok=True)
    with open(SAIDA, "w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, separators=(",", ":"))
    total = sum(r["valor"] for r in registros)
    print(f"\nOK: {len(registros)} registros | {len(set(r['emenda'] for r in registros))} emendas distintas")
    print(f"Valor total: R$ {total:,.2f}")
    print(f"Inconsistências: {n_incons} registro(s) — "
          f"Regra 2.A (Mod. Aplic. ≠ 90): {n_mod} | Regra 2.B (UO × justificativa): {n_uo} | "
          f"alertas revisados e descartados: {len(descartadas)}")
    print(f"JSON gravado em {os.path.abspath(SAIDA)}")


if __name__ == "__main__":
    main()
